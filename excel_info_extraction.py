#!/usr/bin/env python
# coding: utf-8

"""
-----------------------     excel_info_extraction()    -----------------------

Script designed to extract information from cells that are highlighed in 
specific color or edited in the .xlsx file.

The program is run by typing "excel_info_extraction.py" in the program
directory in the Ancaconda console.

The working directory will be asked at the beginning. Then, user needs to type
in what's the source file name and the spreadsheet name. Output will be saved
to the same directory as source file and it is named as OUTPUT.xlsx.

Programmed by Nickole Li 11 November 2020.

"""


# ***********************************************************************
# ***********************     GPLv3 License      ************************
# ***********************************************************************
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
# ***********************************************************************


import os
import tkinter as tk
from tkinter import filedialog as fd
import re
from openpyxl import load_workbook
import pandas as pd


# ---------------------------------------------------------------------------
# Information extraction
# ---------------------------------------------------------------------------
def info_extraction(dir_name, file_name, sheet_name):
    source_file = dir_name + os.sep + file_name
    # Read .xlsx file to wb
    wb = load_workbook(source_file, data_only=True)
    # Read spreadsheet based on the input sheet_name
    sh = wb[sheet_name]

    # Create a blank list for output values
    data = []

    # Working column by column
    for column in sh.columns:
        # Working cell by cell
        for cell in column:
            # ---------------------------------------------------------------
            # Check cell background color
            # ---------------------------------------------------------------
            # Yellow color code is "FFFFFF00" in openpyxl
            if cell.fill.start_color.index == "FFFFFF00": 
                # Check the value inside the cell
                # If it's "Yes", then need to extract info from the next cell in the same column
                if cell.value == "Yes":
                    question_no = str(sh.cell(row=cell.row + 1, column=2).value)
                    question = str(
                        sh.cell(row=cell.row, column=3).value
                        + " "
                        + sh.cell(row=cell.row + 1, column=3).value
                    )
                    text = str(sh.cell(row=cell.row + 1, column=cell.column).value)
                # If the value is not "Yes", extract value from the cell itself
                else:
                    question_no = str(sh.cell(row=cell.row, column=2).value)
                    question = str(sh.cell(row=cell.row, column=3).value)
                    text = str(sh.cell(row=cell.row, column=cell.column).value)
                home = str(sh.cell(row=3, column=cell.column).value)
                office = str(sh.cell(row=14, column=cell.column).value)
                if re.search(r"\|(\s)*", text):
                    text = re.split(r"\|\s*", text, 1)[1]
                response = re.split(r"\s*\[Tag", text, 1)[0]
                mgr_date = re.split(r"Tag\s*\-\s*", text, 1)[1]
                mgr = re.findall(r"[a-zA-Z\& ]+", mgr_date)[0]
                date = re.findall(r"\d+\/\d+\/\d+", mgr_date)[0]
                # Append the extracted info into the "data" list
                data.append(
                    {
                        "Filter": "Filter based on color",
                        "Cell": cell,
                        "Home": home,
                        "Office": office,
                        "Question No": question_no,
                        "Question": question,
                        "Response": response,
                        "Manager": mgr,
                        "Date": date,
                    }
                )
                # Print the worked cell name
                # Can use it as a troubleshooting output
                print(cell)
            # ---------------------------------------------------------------
            # Check string without background color
            # ---------------------------------------------------------------
            # Find the matching flag in cells that were not highlighted
            # in case the edited cells were not highlighted by human errors
            else:
                # Check whether the cell above is marked as yellow,
                # If yes, it means it has been covered before
                if (
                        cell.row > 1
                        and sh.cell(row=cell.row - 1, column=cell.column).fill.start_color.index
                        == "FFFFFF00"
                    ):
                    continue
                else:
                    text = str(sh.cell(row=cell.row, column=cell.column).value)
                    if re.search("RC Survey", text):
                        home = str(sh.cell(row=3, column=cell.column).value)
                        office = str(sh.cell(row=14, column=cell.column).value)
                        question_no = str(sh.cell(row=cell.row, column=2).value)
                        question = str(sh.cell(row=cell.row, column=3).value)
                        text = str(sh.cell(row=cell.row, column=cell.column).value)
                        if re.search(r"\|(\s)*", text):
                            text = re.split(r"\|\s*", text, 1)[1]
                        response = re.split(r"\s*\[Tag", text, 1)[0]
                        mgr_date = re.split(r"Tag\s*\-\s*", text, 1)[1]
                        mgr = re.findall(r"[a-zA-Z\& ]+", mgr_date)[0]
                        date = re.findall(r"\d+\/\d+\/\d+", mgr_date)[0]
                        # Append the extracted info into the "data" list
                        data.append(
                            {
                                "Filter": "Filter based on string match",
                                "Cell": cell,
                                "Home": home,
                                "Office": office,
                                "Question No": question_no,
                                "Question": question,
                                "Response": response,
                                "Manager": mgr,
                                "Date": date,
                            }
                        )
                        # Print the worked cell name
                        # Can use it as a troubleshooting output
                        print(cell)

    # Tranform the list into dataframe
    df = pd.DataFrame(data)
    output_writer = pd.ExcelWriter(
        ROOT.dirname + os.sep + "OUTPUT.xlsx", engine="xlsxwriter",
    )
    # Extract all info into an excel
    df.to_excel(output_writer, index=False)
    # Save file
    output_writer.save()


# ---------------------------------------------------------------------------
# MAIN PROGRAM
# ---------------------------------------------------------------------------

# DIRECTORY DIALOG
ROOT = tk.Tk()
ROOT.withdraw()
ROOT.dirname = fd.askdirectory(
    title="Please select the directory."
    )

# Input the file name
FILE = input("What's the source file name (including file extension)? ")
# Input the spreadsheet name
SHEET = input("What's the spreadsheet name? ")

# Call function info_extraction(dir_name, file_name, sheet_name)
info_extraction(ROOT.dirname, FILE, SHEET)

# Output message
print("Successful!")
